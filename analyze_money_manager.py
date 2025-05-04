import re
from openpyxl import load_workbook
from pprint import pprint
from datetime import datetime
from utils.helpers import write_to_file

def process_coffee_shops(food_data):
    coffee_shops = [
        "Starbucks", "Dunkin Donuts", "Pickup Coffee", "Coffee Project",
        "Tim Hortons", "Dean & Deluca", "Cafe Amazon", "Highlands", "Bo's"
    ]
    coffee_shops_iloilo = ["Coffeebreak", "Cafe Brewtherhood", "Teepee", "Tiring"]

    # adjust depending your coffee shop config here
    coffee_shops = coffee_shops + coffee_shops_iloilo
    
    # filter food_data that matches coffee_shops and is_coffee_shop to true
    food_data_coffee_shops_filtered = [
        (name, data) for name, data in food_data.items() 
        if data[3] and any(re.search(rf'\b{re.escape(coffee_shop)}\b', name, re.IGNORECASE) for coffee_shop in coffee_shops)
    ]
    top_10_coffee_shops = sorted(food_data_coffee_shops_filtered, key=lambda item: item[1][0], reverse=True)[:10]

    return top_10_coffee_shops

def process_fast_foods(food_data):
    fast_foods = ["Jollibee", "Mcdo", "KFC", "Chowking", "Mang Inasal", "Burger King", "Pizza Hut"]
    
    # filter food_data that matches fast_foods
    food_data_fast_foods_filtered = [
        (name, data) for name, data in food_data.items() 
        if any(re.search(rf'\b{re.escape(fast_food)}\b', name, re.IGNORECASE) for fast_food in fast_foods)
    ]
    top_10_fast_foods = sorted(food_data_fast_foods_filtered, key=lambda item: item[1][0], reverse=True)[:10]

    return top_10_fast_foods

def analyze_money_manager(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    header_row  = next(sheet.iter_rows(values_only=True))

    period_index = header_row.index("Period") + 1
    account_index = header_row.index("Accounts") + 1
    category_index = header_row.index("Category") + 1
    subcategory_index = header_row.index("Subcategory") + 1
    amount_index = header_row.index("Amount") + 1
    income_expense_column_index = header_row.index("Income/Expense") + 1
    note_index = header_row.index("Note") + 1

    # totals
    total_expense = 0
    total_expense_count = 0
    total_income = 0
    total_income_count = 0
    
    # data
    expense_account_data = {}
    income_from_data = {}
    purchase_from_data = {}
    food_data = {}
    grocery_data = {}

    # special cases
    total_shopee_count = 0
    total_lazada_count = 0
    total_amazon_count = 0
    total_grab_food_count = 0
    total_grab_car_count = 0
    total_foodpanda_count = 0
    total_711 = 0

    for row in sheet.iter_rows(min_row=2):
        period_cell = row[period_index - 1]
        account_cell = row[account_index - 1]
        category_cell = row[category_index - 1]
        subcategory_cell = row[subcategory_index - 1]
        amount_cell = row[amount_index - 1]
        income_expense_cell = row[income_expense_column_index - 1]
        note_cell = row[note_index - 1]

        # get the values from the cells
        period_value = period_cell.value
        category_value = category_cell.value
        subcategory_value = subcategory_cell.value
        amount_value = amount_cell.value
        income_expense_value = income_expense_cell.value
        account_value = account_cell.value
        note_value = note_cell.value

        if isinstance(amount_value, (int, float)):
            if income_expense_value == "Exp.":
                total_expense_count += 1
                total_expense += amount_value

                # count expense accounts
                if account_value:
                    expense_account_data[account_value] = expense_account_data.get(account_value, 0) + 1

                if note_value:
                    # count purchase_from entry, total amount, and first instance
                    if note_value not in purchase_from_data:
                        purchase_from_data[note_value] = (0, 0, period_value.strftime("%B %d, %Y"))
                    current_count, current_amount_total, current_period_value = purchase_from_data[note_value]

                    # check if current_period_value is older than period_value then update it
                    if datetime.strptime(current_period_value, "%B %d, %Y") > period_value:
                        current_period_value = period_value.strftime("%B %d, %Y")
                    purchase_from_data[note_value] = (current_count + 1, round(current_amount_total + amount_value, 2), current_period_value)

                    if category_value == "Food":
                        # include Grab and Foodpanda for delivery
                        is_coffee_shop = subcategory_value == "Cafe Hopping" or subcategory_value == "Grab" or subcategory_value == "Foodpanda"

                        if note_value not in food_data:
                            food_data[note_value] = (0, 0, period_value.strftime("%B %d, %Y"), is_coffee_shop)
                        current_count, current_amount_total, current_period_value, _ = food_data[note_value]

                        # check if current_period_value is older than period_value then update it
                        if datetime.strptime(current_period_value, "%B %d, %Y") > period_value:
                            current_period_value = period_value.strftime("%B %d, %Y")
                        food_data[note_value] = (current_count + 1, round(current_amount_total + amount_value, 2), current_period_value, is_coffee_shop)

                        # catch GrabFood
                        if subcategory_value == "Grab":
                            total_grab_food_count += 1
                        
                        # catch Foodpanda
                        if subcategory_value == "Foodpanda":
                            total_foodpanda_count += 1
                    
                    elif category_value == "Grocery":
                        if note_value not in grocery_data:
                            grocery_data[note_value] = (0, 0, period_value.strftime("%B %d, %Y"))
                        current_count, current_amount_total, current_period_value = grocery_data[note_value]
                        # check if current_period_value is older than period_value then update it
                        if datetime.strptime(current_period_value, "%B %d, %Y") > period_value:
                            current_period_value = period_value.strftime("%B %d, %Y")
                        grocery_data[note_value] = (current_count + 1, round(current_amount_total + amount_value, 2), current_period_value)

                    elif category_value == "Transportation":
                        # catch grabcar and grabtaxi
                        if re.search(r'^Grab', note_value, re.IGNORECASE):
                            total_grab_car_count += 1
                        
                    # catch shopee orders
                    if re.search(r'Shopee$', note_value, re.IGNORECASE):
                        total_shopee_count += 1

                    # catch lazada orders
                    if re.search(r'Lazada$', note_value, re.IGNORECASE):
                        total_lazada_count += 1

                    # catch amazon orders
                    if re.search(r'Amazon$', note_value, re.IGNORECASE):
                        total_amazon_count += 1

                    # catch 711 orders
                    if re.search(r'^711', note_value, re.IGNORECASE):
                        total_711 += 1
                            
                
            elif income_expense_value == "Income":
                total_income_count += 1
                total_income += amount_value

                if note_value:
                    # count income_from entry, total amount, and first instance
                    if note_value not in income_from_data:
                        income_from_data[note_value] = (0, 0, period_value.strftime("%B %d, %Y"))
                    current_count, current_amount_total, current_period_value = income_from_data[note_value]

                    # check if current_period_value is older than period_value then update it
                    if datetime.strptime(current_period_value, "%B %d, %Y") > period_value:
                        current_period_value = period_value.strftime("%B %d, %Y")
                    income_from_data[note_value] = (current_count + 1, round(current_amount_total + amount_value, 2), current_period_value)

    output = []
    output.append("# Money Manager Analysis")
    output.append("@crrmacarse")

    output.append("\n## Summary")
    output.append(f"- Total Income: PHP {total_income:,.2f}")
    output.append(f"- Total Expense: PHP {total_expense:,.2f}")
    balance = total_income - total_expense
    output.append(f"- Balance: PHP {balance:,.2f}")
    output.append(f"- Total Income Entry: {total_income_count}")
    output.append(f"- Total Expense Entry: {total_expense_count}")

    # Most common category, subcategory

    output.append("\n## Expense Accounts")
    output.append("| Account | Number of Entries ↓ |")
    output.append("|-------------------|-------------------|")
    sorted_expense_accounts = sorted(expense_account_data.items(), key=lambda item: item[1], reverse=True)
    for account, count in sorted_expense_accounts:
        output.append(f"| {account} | {count} |")

    output.append("\n## Top 10 Income From")
    output.append("| Income from | Number of Entries | Total Amount ↓ | First Instance   |")
    output.append("|-------------------|-------------------|--------------|------------------|")
    top_10_income_from_data = sorted(income_from_data.items(), key=lambda item: item[1][1], reverse=True)[:10]
    for income_from, (count, total, first_instance) in top_10_income_from_data:
        output.append(f"| {income_from} | {count} | PHP {total:,.2f} | {first_instance} |")

    # Top 5 Highest income earned with date

    output.append("\n## Top 30 Expense From")
    top_30_purchase_from_data = sorted(purchase_from_data.items(), key=lambda item: item[1][0], reverse=True)[:30]
    output.append("| Expense from | Number of Entries ↓ | Total Amount | First Instance   |")
    output.append("|-------------------|-------------------|--------------|------------------|")
    for purchase_from, (count, total, first_instance) in top_30_purchase_from_data:
        output.append(f"| {purchase_from} | {count} | PHP {total:,.2f} | {first_instance} |")

    output.append("\n## Top 10 Expense From by Amount")
    top_30_purchase_from_data = sorted(purchase_from_data.items(), key=lambda item: item[1][1], reverse=True)[:10]
    output.append("| Expense from | Total Amount ↓ |")
    output.append("|-------------------|-------------------|")
    for purchase_from, (_, total, _) in top_30_purchase_from_data:
        output.append(f"| {purchase_from} | PHP {total:,.2f} |")

    # Top 10 Most expensive expense with date

    output.append("\n## Top 10 Grocery Store")
    output.append("| Grocery | Number of Entries ↓ | Total Amount | First Instance   |")
    output.append("|-------------------|-------------------|--------------|------------------|")
    top_10_grocery_data = sorted(grocery_data.items(), key=lambda item: item[1][0], reverse=True)[:10]
    for grocery, (count, total, first_instance) in top_10_grocery_data:
        output.append(f"| {grocery} | {count} | PHP {total:,.2f} | {first_instance} |")

    output.append("\n## Top 10 Grocery Store by Amount")
    top_10_grocery_data = sorted(grocery_data.items(), key=lambda item: item[1][1], reverse=True)[:10]
    output.append("| Grocery | Total Amount ↓ |")
    output.append("|-------------------|-------------------|")
    for grocery, (_, total, _) in top_10_grocery_data:
        output.append(f"| {grocery} | PHP {total:,.2f} |")

    output.append("\n## Top 30 Food")
    top_30_food_data = sorted(food_data.items(), key=lambda item: item[1][0], reverse=True)[:30]
    output.append("| Food Establishments | Number of Entries ↓ | Total Amount | First Instance |")
    output.append("|-------------------|-------------------|--------------|------------------|") 
    for food_establishment, (count, total, first_instance, _) in top_30_food_data:
        output.append(f"| {food_establishment} | {count} | PHP {total:,.2f} | {first_instance} |")

    output.append("\n## Top 10 Fast Foods")
    top_10_fast_foods = process_fast_foods(food_data)
    output.append("| Fast Food | Number of Entries ↓ | Total Amount |")
    output.append("|-------------------|-------------------|-------------------|")
    for fast_food, (count, total, _, _) in top_10_fast_foods:
        output.append(f"| {fast_food} | {count} | PHP {total:,.2f} |")

    output.append("\n## Top 10 Coffee Shops")
    top_10_coffee_shops = process_coffee_shops(food_data)
    output.append("| Coffee Shop | Number of Entries ↓ | Total Amount |")
    output.append("|-------------------|-------------------|-------------------|")
    for coffee_shop, (count, total, _, _) in top_10_coffee_shops:
        output.append(f"| {coffee_shop} | {count} | PHP {total:,.2f} |")

    output.append("\n## Special Cases")
    output.append(f"- Total Shopee Order Count: {total_shopee_count}")
    output.append(f"- Total Lazada Order Count: {total_lazada_count}")
    output.append(f"- Total Amazon Order Count: {total_amazon_count}")
    output.append(f"- Total Foodpanda Count: {total_foodpanda_count}")
    output.append(f"- Total GrabFood Count: {total_grab_food_count}")
    output.append(f"- Total GrabCar Count: {total_grab_car_count}")
    output.append(f"- Total 711 Count: {total_711}")

    return output

if __name__ == "__main__":
    try:
        # temporarily added a default file path
        file_path = input("Enter the path to the .xlsx file: ") or "./dump/mm.xlsx"
        output_file_path = input("Enter the path to the output .md file (default: dump/output.md): ") or "dump/output.md"

        result = analyze_money_manager(file_path)

        write_to_file(output_file_path, "\n".join(result))
    except Exception as e:
        print(f"Error: {e}")