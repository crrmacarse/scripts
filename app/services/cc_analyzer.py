"""CC Analyzer service.

Refactored from cc_analyzer_plus.py to be web-friendly:
- Reads PDF from in-memory bytes (file upload), not from a path on disk.
- Reads Money Manager data from either the Google Sheet (default) or an
  uploaded xlsx export.
- Returns a structured analysis (matched / missing / unused / inaccurate) so
  the UI can flag issues before any write to the destination sheet.
- Keeps the original Google Sheets write behaviour for the confirm step.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Iterable, Literal

from PyPDF2 import PdfReader
from dateutil.relativedelta import relativedelta
from gspread_formatting import CellFormat, TextFormat, format_cell_range

from app.services.google_auth import get_gspread_client


RowStatus = Literal["matched", "missing", "inaccurate", "duplicate"]


@dataclass
class MMEntry:
    date: str  # m/d/yy — same as raw_period date but in 2-digit-year form,
               # used internally for matching against PDF tran dates
    amount: str  # "1,234.56"
    description: str
    raw_period: str  # m/d/Y — user-facing date, shifted -1 day from the MM
                     # sheet export so it matches what the MM app shows
    category: str = ""
    subcategory: str = ""
    note: str = ""
    in_cycle: bool = True  # False = within the ±3 day fetch extension but
                           # outside the strict billing cycle. Excluded from
                           # auto-matching; hidden in the MM table by default.

    @property
    def amount_value(self) -> float:
        return float(self.amount.replace(",", ""))

    @property
    def expense_name(self) -> str:
        if self.category and self.subcategory:
            base = f"{self.category} > {self.subcategory}"
        else:
            base = self.category or self.subcategory or ""
        if self.note:
            return f"{base} · {self.note}" if base else self.note
        return base


@dataclass
class PdfRow:
    tran_date: str
    post_date: str
    description: str
    amount: str  # "1,234.56"

    @property
    def amount_value(self) -> float:
        return float(self.amount.replace(",", ""))


@dataclass
class AnalyzedRow:
    pdf: PdfRow
    status: RowStatus
    matched_mm: MMEntry | None = None
    nearest_candidates: list[MMEntry] = field(default_factory=list)
    candidate_unused_ids: list[int] = field(default_factory=list)
    note: str = ""
    mm_mentions: str = ""
    row_id: str = ""


@dataclass(frozen=True)
class CreditCard:
    key: str
    label: str
    mm_account: str
    sheet_name: str
    due_days: int = 21  # calendar days from cutoff to due date


CREDIT_CARDS: list[CreditCard] = [
    CreditCard(
        key="security-bank-world",
        label="Security Bank World",
        mm_account="Security Bank World",
        sheet_name="Security Bank World CC",
        due_days=21,
    ),
]


def get_credit_card(key: str) -> CreditCard:
    for card in CREDIT_CARDS:
        if card.key == key:
            return card
    raise ValueError(f"Unknown credit card: {key!r}")


@dataclass
class AnalysisResult:
    sheet_name: str
    billing_period: str
    cutoff_date: str  # ISO yyyy-mm-dd
    due_date: str  # human-readable, weekend-adjusted
    due_date_iso: str  # yyyy-mm-dd
    rows: list[AnalyzedRow]
    unused_mm: list[MMEntry]
    total_amount: float
    total_cr: float
    mm_total_count: int = 0
    mm_total_amount: float = 0.0

    @property
    def grand_total(self) -> float:
        return self.total_amount - self.total_cr

    @property
    def cc_statement_total(self) -> float:
        """What you owe the bank for this cycle: debits − CR refunds."""
        return self.grand_total

    @property
    def cc_vs_mm_difference(self) -> float:
        """MM minus CC. Positive = MM has more than CC (excess to remove).
        Negative = MM short of CC (entries missing in MM). Zero = reconciled."""
        return round(self.mm_total_amount - self.cc_statement_total, 2)

    @property
    def issue_count(self) -> int:
        return sum(1 for r in self.rows if r.status != "matched")

    def serialise_for_confirm(self) -> dict:
        return {
            "sheet_name": self.sheet_name,
            "billing_period": self.billing_period,
            "cutoff_date": self.cutoff_date,
            "total_amount": self.total_amount,
            "total_cr": self.total_cr,
            "rows": [
                {
                    "tran_date": r.pdf.tran_date,
                    "post_date": r.pdf.post_date,
                    "description": r.pdf.description,
                    "amount": r.pdf.amount,
                    # Drop algorithm hint prefixes; the user verified before
                    # pushing so [?] / [!] notes are noise on the sheet.
                    "note": "" if r.note.startswith(("[?]", "[!]")) else r.note,
                    "mm_mentions": r.mm_mentions,
                }
                for r in self.rows
            ],
        }


_TYPE_EXPENSE = "Exp."
# Days of slack on each side of the strict billing cycle when fetching MM
# entries — exposed via the "Show all MM data" toggle so the user can spot
# drift cases just outside the cycle boundary.
_MM_EXTENSION_DAYS = 3

_MENTION_RE = re.compile(r"@\w+")
_PDF_TABLE_RE = re.compile(
    r"(\d{2}/\d{2}/\d{2})\s+(\d{2}/\d{2}/\d{2})\s+(.+?)\s+([\d,]+\.\d{2})(?:\s+(CR))?"
)


def _extract_mentions(description: str) -> str:
    return ", ".join(_MENTION_RE.findall(description or ""))


def _strip_mentions(description: str) -> str:
    return _MENTION_RE.sub("", description or "").strip()


def _parse_date_loose(value: str) -> datetime:
    for fmt in ("%m/%d/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    raise ValueError(f"Unrecognised date: {value!r}")


def _sort_date_key(value: str) -> datetime:
    """`_parse_date_loose` with a sentinel fallback so unparseable values
    sort to the end of an ascending list instead of raising."""
    try:
        return _parse_date_loose(value)
    except ValueError:
        return datetime.max


def _money_str(value: float) -> str:
    return f"{value:,.2f}"


# --- Money Manager loaders ---------------------------------------------------

def adjust_to_banking_day(d: datetime) -> datetime:
    """Bump a date forward to the next weekday (Mon–Fri).

    Holidays are not modelled — if the resulting weekday is a public holiday
    the actual banking day will be later still. The form/result pages note
    this caveat for the user.
    """
    while d.weekday() >= 5:  # 5=Sat, 6=Sun
        d += timedelta(days=1)
    return d


def compute_due_date(cutoff_date: datetime, card: CreditCard) -> datetime:
    return adjust_to_banking_day(cutoff_date + timedelta(days=card.due_days))


def compute_billing_period(cutoff_date: datetime) -> tuple[datetime, datetime]:
    """The billing cycle that the bill cut on `cutoff_date` covers.

    Convention: the cutoff date itself sits at the start of the next cycle;
    this bill's window is (cutoff − 1 month) … (cutoff − 1 day).
    For cutoff 4/20 → 3/20 … 4/19.
    """
    start = cutoff_date - relativedelta(months=1)
    end = cutoff_date - timedelta(days=1)
    return (start, end)


def compute_mm_window(cutoff_date: datetime) -> tuple[datetime, datetime]:
    """MM lookup window = strict billing cycle, no padding.

    Anchored to the cutoff: `(cutoff − 1 month + 1 day) … cutoff`. Entries
    outside this window aren't fetched, so they never appear in the MM table
    or as candidates.
    """
    return compute_billing_period(cutoff_date)


def load_mm_from_google_sheet(
    window_start: datetime,
    window_end: datetime,
    mm_account: str,
) -> list[MMEntry]:
    client = get_gspread_client()
    sheet = client.open("Money Manager Data")
    worksheet = sheet.worksheet("Sheet1")
    data = worksheet.get_all_values()
    return _parse_mm_rows(data, window_start, window_end, mm_account)


def load_mm_from_xlsx(
    file_bytes: bytes,
    window_start: datetime,
    window_end: datetime,
    mm_account: str,
) -> list[MMEntry]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    data = [list(map(_xlsx_cell_to_str, row)) for row in ws.iter_rows(values_only=True)]
    return _parse_mm_rows(data, window_start, window_end, mm_account)


def _xlsx_cell_to_str(cell) -> str:
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        return cell.strftime("%m/%d/%Y")
    return str(cell)


def _parse_mm_rows(
    data: list[list[str]],
    window_start: datetime,
    window_end: datetime,
    mm_account: str,
) -> list[MMEntry]:
    if not data:
        raise ValueError("Money Manager data appears empty.")

    header = data[0]
    try:
        period_idx = header.index("Period")
        accounts_idx = header.index("Accounts")
        amount_idx = header.index("Amount")
        type_idx = header.index("Income/Expense")
        description_idx = header.index("Description")
    except ValueError as exc:
        raise ValueError(f"Missing column in Money Manager data: {exc}") from exc

    category_idx = header.index("Category") if "Category" in header else None
    subcategory_idx = header.index("Subcategory") if "Subcategory" in header else None
    note_idx = header.index("Note") if "Note" in header else None

    base_required = max(period_idx, amount_idx, accounts_idx, type_idx, description_idx)

    entries: list[MMEntry] = []
    for row in data[1:]:
        if len(row) <= base_required:
            continue
        if row[accounts_idx] != mm_account:
            continue
        if row[type_idx] != _TYPE_EXPENSE:
            continue

        try:
            period_dt = _parse_date_loose(row[period_idx])
        except ValueError:
            continue

        # Fetch a small ±3 day extension around the strict cycle so that
        # drifted entries near the boundary are still available for the
        # user to inspect via the "Show all MM data" toggle. Each entry
        # carries an `in_cycle` flag — strictly-inside entries are part of
        # the auto-matching pool and visible by default; outside-but-near
        # entries are hidden by default and skipped by the matcher.
        extension_start = window_start - timedelta(days=_MM_EXTENSION_DAYS)
        extension_end = window_end + timedelta(days=_MM_EXTENSION_DAYS)
        if not (extension_start <= period_dt <= extension_end):
            continue
        in_cycle = window_start <= period_dt <= window_end

        adjusted = period_dt.strftime("%m/%d/%y")
        try:
            amount_value = float(str(row[amount_idx]).replace(",", ""))
        except ValueError:
            continue

        category = row[category_idx] if category_idx is not None and len(row) > category_idx else ""
        subcategory = (
            row[subcategory_idx] if subcategory_idx is not None and len(row) > subcategory_idx else ""
        )
        note = row[note_idx] if note_idx is not None and len(row) > note_idx else ""

        entries.append(
            MMEntry(
                date=adjusted,
                amount=_money_str(amount_value),
                description=row[description_idx] or "",
                raw_period=period_dt.strftime("%m/%d/%Y"),
                category=category or "",
                subcategory=subcategory or "",
                note=note or "",
                in_cycle=in_cycle,
            )
        )
    return entries


# --- PDF extraction ----------------------------------------------------------

@dataclass
class PdfExtraction:
    rows: list[PdfRow]
    total_amount: float
    total_cr: float


def extract_pdf(file_bytes: bytes, password: str | None) -> PdfExtraction:
    reader = PdfReader(io.BytesIO(file_bytes))
    if reader.is_encrypted:
        if not password:
            raise ValueError("PDF is password protected. Provide a password.")
        try:
            ok = reader.decrypt(password)
        except Exception as exc:
            raise ValueError("Could not decrypt PDF.") from exc
        if not ok:
            raise ValueError("Incorrect PDF password.")

    rows: list[PdfRow] = []
    total_amount = 0.0
    total_cr = 0.0

    for page in reader.pages:
        text = page.extract_text() or ""
        for tran_date, post_date, description, amount, is_cr in _PDF_TABLE_RE.findall(text):
            amount_clean = amount.replace(",", "")
            try:
                amount_value = float(amount_clean)
            except ValueError:
                continue

            if is_cr:
                if not description.startswith("PAYMENT"):
                    total_cr += amount_value
                continue

            total_amount += amount_value
            rows.append(
                PdfRow(
                    tran_date=tran_date,
                    post_date=post_date,
                    description=description.strip(),
                    amount=_money_str(amount_value),
                )
            )

    return PdfExtraction(rows=rows, total_amount=total_amount, total_cr=total_cr)


# --- Matching / analysis ----------------------------------------------------

# Within the period of the CC bill we scan the *whole* unmatched pool
# (no per-tx day window) and rank by nearest date, with amount drift as the
# tiebreaker. The matching window is bounded by the PDF's actual transaction
# date range (with a small pad for MM Period drift) so we don't pull entries
# from outside the bill's period when picking automatic matches. The user
# can still manually link to anything in the full cycle pool via the dropdown.
_INACCURATE_AMOUNT_TOLERANCE = 0.05  # 5%
_MATCH_PERIOD_PAD_DAYS = 3
# Token-overlap ratio for the merchant-only candidate path. 0.7 avoids
# false positives where only generic location tokens (e.g. "taiwan" /
# "kaohsiung") overlap between the CC merchant and an unrelated MM entry.
_MERCHANT_MATCH_THRESHOLD = 0.7
# Even a strong merchant hit shouldn't override a wildly different amount.
# Cap merchant-only inclusion at ±50% drift — typos within ~1.5x slip
# through, but a 173 vs 2,729 mismatch (≈1475% drift) does not.
_MERCHANT_ONLY_MAX_DRIFT = 0.5

_MERCHANT_TOKEN_RE = re.compile(r"[a-z0-9]+")
_MERCHANT_NOISE_TOKENS: frozenset[str] = frozenset({
    # generic corporate suffixes
    "inc", "corp", "co", "ltd", "llc", "lp", "intl", "international",
    # geo / locality words common in CC merchant strings
    "phl", "ph", "philippines", "manila", "makati", "qc", "quezon",
    "city", "metro", "mnl", "bgc", "ortigas", "pasig", "taguig",
    # filler
    "the", "and", "of", "for",
})


def _merchant_tokens(s: str) -> set[str]:
    return {
        t for t in _MERCHANT_TOKEN_RE.findall((s or "").lower())
        if len(t) >= 3 and t not in _MERCHANT_NOISE_TOKENS
    }


def _merchant_score(merchant: str, mm_entry: "MMEntry") -> float:
    """Token-overlap similarity between the CC merchant string and the MM
    entry's expense name + description. 0.0 = no signal, 1.0 = full overlap.
    """
    m_tokens = _merchant_tokens(merchant)
    if not m_tokens:
        return 0.0
    e_text = (mm_entry.expense_name or "") + " " + (mm_entry.description or "")
    e_tokens = _merchant_tokens(e_text)
    if not e_tokens:
        return 0.0
    overlap = m_tokens & e_tokens
    if not overlap:
        return 0.0
    return len(overlap) / min(len(m_tokens), len(e_tokens))


def _within(mm_date: datetime, start: datetime | None, end: datetime | None) -> bool:
    if start and mm_date < start:
        return False
    if end and mm_date > end:
        return False
    return True


def _exact_match(
    pdf_row: PdfRow,
    mm_pool: list[MMEntry],
    *,
    period_start: datetime | None = None,
    period_end: datetime | None = None,
) -> MMEntry | None:
    """Best-fit entry inside the period whose amount matches exactly.
    Ranking: highest merchant-name similarity first, then nearest date."""
    pdf_date = _parse_date_loose(pdf_row.tran_date)
    scored: list[tuple[float, int, MMEntry]] = []
    for entry in mm_pool:
        if not entry.in_cycle:
            continue
        if entry.amount != pdf_row.amount:
            continue
        try:
            mm_date = _parse_date_loose(entry.date)
        except ValueError:
            continue
        if not _within(mm_date, period_start, period_end):
            continue
        score = _merchant_score(pdf_row.description, entry)
        # Negate score so ascending sort puts the strongest match first.
        scored.append((-score, abs((pdf_date - mm_date).days), entry))
    if not scored:
        return None
    scored.sort(key=lambda t: (t[0], t[1]))
    return scored[0][2]


def _nearby_candidates(
    pdf_row: PdfRow,
    mm_pool: Iterable[MMEntry],
    *,
    period_start: datetime | None = None,
    period_end: datetime | None = None,
) -> list[MMEntry]:
    """Inaccurate candidates inside the period.

    Inclusion: amount within ±5% OR merchant-name similarity ≥ threshold.
    The merchant-only path catches rows where the user typed the wrong
    amount in MM but the merchant clearly matches the CC tx.

    Sort: earliest-by-date first, matching the "all date sorts earliest
    first" rule applied to other display lists.
    """
    pdf_amount = pdf_row.amount_value
    if pdf_amount == 0:
        return []
    scored: list[tuple[datetime, MMEntry]] = []
    for entry in mm_pool:
        if not entry.in_cycle:
            continue
        try:
            mm_date = _parse_date_loose(entry.date)
        except ValueError:
            continue
        if not _within(mm_date, period_start, period_end):
            continue
        amount_drift = abs(entry.amount_value - pdf_amount) / pdf_amount
        merchant = _merchant_score(pdf_row.description, entry)
        # Inclusion paths:
        #   (a) close-on-amount    — drift ≤ 5%
        #   (b) merchant-only path — merchant ≥ 0.7 AND drift ≤ 50%
        within_amount = amount_drift <= _INACCURATE_AMOUNT_TOLERANCE
        merchant_strong = (
            merchant >= _MERCHANT_MATCH_THRESHOLD
            and amount_drift <= _MERCHANT_ONLY_MAX_DRIFT
        )
        if not (within_amount or merchant_strong):
            continue
        scored.append((mm_date, entry))
    scored.sort(key=lambda t: t[0])
    return [e for _, e in scored]


def _cc_period_bounds(
    pdf_extraction: PdfExtraction,
) -> tuple[datetime | None, datetime | None]:
    """Date range of the PDF's transactions, padded ±_MATCH_PERIOD_PAD_DAYS
    to absorb MM Period drift."""
    dates: list[datetime] = []
    for r in pdf_extraction.rows:
        try:
            dates.append(_parse_date_loose(r.tran_date))
        except ValueError:
            continue
    if not dates:
        return (None, None)
    pad = timedelta(days=_MATCH_PERIOD_PAD_DAYS)
    return (min(dates) - pad, max(dates) + pad)


def analyse(
    *,
    pdf_extraction: PdfExtraction,
    mm_entries: list[MMEntry],
    sheet_name: str,
    billing_period: str,
    cutoff_date: datetime,
    card: CreditCard,
) -> AnalysisResult:
    mm_pool = list(mm_entries)
    analysed: list[AnalyzedRow] = []
    period_start, period_end = _cc_period_bounds(pdf_extraction)

    # All date-bearing lists are sorted earliest-first for consistent display.
    sorted_pdf_rows = sorted(pdf_extraction.rows, key=lambda r: _sort_date_key(r.tran_date))

    for pdf_row in sorted_pdf_rows:
        match = _exact_match(
            pdf_row, mm_pool, period_start=period_start, period_end=period_end
        )
        if match is not None:
            mm_pool.remove(match)
            note = _strip_mentions(match.description)
            mentions = _extract_mentions(match.description)
            analysed.append(
                AnalyzedRow(
                    pdf=pdf_row,
                    status="matched",
                    matched_mm=match,
                    note=note,
                    mm_mentions=mentions,
                )
            )
            continue

        candidates = _nearby_candidates(
            pdf_row, mm_pool, period_start=period_start, period_end=period_end
        )
        if candidates:
            # Don't consume the candidate from mm_pool — leave it in `unused`
            # so the UI can show the linkage and the user can resolve the
            # inaccurate row + the unused MM entry together.
            analysed.append(
                AnalyzedRow(
                    pdf=pdf_row,
                    status="inaccurate",
                    nearest_candidates=candidates,
                    note="[?] " + "; ".join(
                        f"MM {c.date} {c.amount} ({_strip_mentions(c.description)})"
                        for c in candidates
                    ),
                )
            )
            continue

        analysed.append(
            AnalyzedRow(pdf=pdf_row, status="missing", note="[!] not found in MM")
        )

    # Per-entry display-date correction for the MM Google Sheet's +1 drift.
    # The export consistently lands an entry's date one day ahead of what
    # the MM app shows for most users, so we DEFAULT to shifting display
    # back by one day. We only keep the sheet date when the PDF positively
    # confirms it: i.e. there's a CC tx with the same amount on the sheet
    # date AND none on the day before. The shift updates both the matching
    # `date` field and the user-facing `raw_period` so a drifted entry
    # exact-matches its CC row instead of appearing 1 day off.
    pdf_amount_by_date: set[tuple[str, str]] = set()
    for r in pdf_extraction.rows:
        try:
            d = _parse_date_loose(r.tran_date).strftime("%Y-%m-%d")
        except ValueError:
            continue
        pdf_amount_by_date.add((r.amount, d))

    shift_cache: dict[int, MMEntry] = {}

    def _shifted(entry: MMEntry) -> MMEntry:
        key = id(entry)
        if key in shift_cache:
            return shift_cache[key]
        try:
            cur = _parse_date_loose(entry.date)
        except ValueError:
            shift_cache[key] = entry
            return entry
        prev = cur - timedelta(days=1)
        cur_iso = cur.strftime("%Y-%m-%d")
        prev_iso = prev.strftime("%Y-%m-%d")
        cur_match = (entry.amount, cur_iso) in pdf_amount_by_date
        prev_match = (entry.amount, prev_iso) in pdf_amount_by_date
        # Keep the sheet date when the PDF clearly anchors it on `cur`;
        # otherwise assume +1 drift and shift back. Default-shift covers
        # unused entries (no PDF correlation) too.
        if cur_match and not prev_match:
            shift_cache[key] = entry
            return entry
        shifted = MMEntry(
            date=prev.strftime("%m/%d/%y"),
            amount=entry.amount,
            description=entry.description,
            raw_period=prev.strftime("%m/%d/%Y"),
            category=entry.category,
            subcategory=entry.subcategory,
            note=entry.note,
            in_cycle=entry.in_cycle,
        )
        shift_cache[key] = shifted
        return shifted

    mm_pool[:] = [_shifted(e) for e in mm_pool]
    for row in analysed:
        if row.matched_mm is not None:
            row.matched_mm = _shifted(row.matched_mm)
        if row.nearest_candidates:
            row.nearest_candidates = [_shifted(c) for c in row.nearest_candidates]

    # Date-sort everything earliest-first for the display lists.
    mm_pool.sort(key=lambda e: _sort_date_key(e.date))
    for row in analysed:
        if row.nearest_candidates:
            row.nearest_candidates.sort(key=lambda c: _sort_date_key(c.date))

    # Stable row ids for the UI, plus unused-id back-references for the
    # inaccurate rows so a checkbox toggle can hide the picked MM entry.
    for i, row in enumerate(analysed):
        row.row_id = f"r{i}"
    unused_index = {id(entry): i for i, entry in enumerate(mm_pool)}
    for row in analysed:
        if row.status == "inaccurate":
            row.candidate_unused_ids = [
                unused_index[id(c)] for c in row.nearest_candidates if id(c) in unused_index
            ]

    due = compute_due_date(cutoff_date, card)
    in_cycle_entries = [e for e in mm_entries if e.in_cycle]
    mm_total_amount = round(sum(e.amount_value for e in in_cycle_entries), 2)
    return AnalysisResult(
        sheet_name=sheet_name,
        billing_period=billing_period,
        cutoff_date=cutoff_date.strftime("%Y-%m-%d"),
        due_date=due.strftime("%a, %b %d %Y"),
        due_date_iso=due.strftime("%Y-%m-%d"),
        rows=analysed,
        unused_mm=mm_pool,
        total_amount=pdf_extraction.total_amount,
        total_cr=pdf_extraction.total_cr,
        mm_total_count=len(in_cycle_entries),
        mm_total_amount=mm_total_amount,
    )


# --- Sheet writer ------------------------------------------------------------


class WorksheetExistsError(Exception):
    """Raised by push_to_google_sheet when the destination worksheet for the
    billing period already exists and force_overwrite was not passed."""


def push_to_google_sheet(payload: dict, *, force_overwrite: bool = False) -> str:
    """Write the analysed rows to a new tab on the destination sheet.

    `payload` matches AnalysisResult.serialise_for_confirm() plus an optional
    `shoulders` key — a list of [name, amount] pairs to render in a small
    side-table to the right of the main data.

    If a worksheet with the same billing-period title already exists,
    `WorksheetExistsError` is raised unless `force_overwrite=True`, in which
    case the existing tab is deleted and replaced.

    Returns the spreadsheet URL.
    """
    client = get_gspread_client()
    sheet = client.open(payload["sheet_name"])

    billing_period = payload["billing_period"]
    existing = next((ws for ws in sheet.worksheets() if ws.title == billing_period), None)
    if existing is not None:
        if not force_overwrite:
            raise WorksheetExistsError(
                f"The billing period worksheet '{billing_period}' already exists."
            )
        sheet.del_worksheet(existing)

    rows = payload["rows"]
    new_ws = sheet.add_worksheet(title=billing_period, rows=str(len(rows) + 10), cols="20")

    worksheets = sheet.worksheets()
    sheet.reorder_worksheets([new_ws] + [ws for ws in worksheets if ws != new_ws])

    new_ws.append_row(
        ["Transaction", "Post date", "Merchant", "Amount", "Notes", "Shoulder", "C", "S"]
    )
    format_cell_range(new_ws, "1:1", CellFormat(textFormat=TextFormat(bold=True)))

    body = [
        [r["tran_date"], r["post_date"], r["description"], r["amount"], r["note"], r["mm_mentions"], "", ""]
        for r in rows
    ]
    body.sort(key=lambda x: _parse_date_loose(x[1]))

    if body:
        new_ws.append_rows(body, value_input_option="USER_ENTERED")

    total_amount = float(payload["total_amount"])
    total_cr = float(payload["total_cr"])
    grand_total = total_amount - total_cr

    new_ws.append_row(["", "", "TOTAL:", _money_str(total_amount)], value_input_option="USER_ENTERED")
    new_ws.append_row(["", "", "REIMBURSED TOTAL:", _money_str(total_cr)], value_input_option="USER_ENTERED")
    new_ws.append_row(["", "", "Grand Total:", _money_str(grand_total)], value_input_option="USER_ENTERED")

    # Shoulder summary side-table at columns J/K. Name + Total only.
    shoulders = payload.get("shoulders") or []
    if shoulders:
        side_values = [["Shoulder", "Total"]]
        for name, amount in shoulders:
            try:
                amt = float(amount)
            except (TypeError, ValueError):
                continue
            side_values.append([str(name), _money_str(amt)])
        side_range = f"J1:K{len(side_values)}"
        new_ws.update(side_range, side_values, value_input_option="USER_ENTERED")
        format_cell_range(new_ws, "J1:K1", CellFormat(textFormat=TextFormat(bold=True)))

    return f"https://docs.google.com/spreadsheets/d/{sheet.id}"
