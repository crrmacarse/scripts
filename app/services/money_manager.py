"""Money Manager analyzer service.

Refactored from analyze_money_manager.py to be web-friendly:
- Reads xlsx from in-memory bytes (file upload), not from a path on disk.
- Returns a structured AnalysisResult so templates can render charts.
- Enriches the original analysis with monthly trend + per-category totals
  (the CLI version only emitted top-N tables; charts need time-series and
  category breakdowns too).

Expected xlsx shape (header row 1, exported from Money Manager iOS/Android):
    Period | Accounts | Category | Subcategory | Amount | Income/Expense | Note | Description

Income/Expense values: "Exp." for expense, "Income" for income.
Description supports @Name tags for shoulder-splitting.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Iterable

from openpyxl import load_workbook


COFFEE_SHOPS_GLOBAL = [
    "Starbucks", "Dunkin Donuts", "Pickup Coffee", "Coffee Project",
    "Tim Hortons", "Dean & Deluca", "Cafe Amazon", "Highlands", "Bo's",
]
COFFEE_SHOPS_ILOILO = ["Coffeebreak", "Cafe Brewtherhood", "Teepee", "Tiring"]
COFFEE_SHOPS = COFFEE_SHOPS_GLOBAL + COFFEE_SHOPS_ILOILO

FAST_FOODS = [
    "Jollibee", "Mcdo", "KFC", "Chowking", "Mang Inasal",
    "Burger King", "Pizza Hut",
]

EXPECTED_COLUMNS = [
    "Period", "Accounts", "Category", "Subcategory", "Amount",
    "Income/Expense", "Note", "Description",
]


@dataclass
class TopEntry:
    name: str
    count: int
    total: float
    first_instance: str = ""


@dataclass
class MonthlyPoint:
    month: str  # "2025-01" (ISO-ish, sorts naturally)
    label: str  # "Jan 2025" (display)
    income: float
    expense: float

    @property
    def net(self) -> float:
        return round(self.income - self.expense, 2)


@dataclass
class TaggedMention:
    name: str  # original casing as first seen, e.g. "Christian"
    count: int
    amount: float


@dataclass
class AnalysisResult:
    total_income: float
    total_expense: float
    balance: float
    total_income_count: int
    total_expense_count: int

    expense_accounts: list[tuple[str, int]]
    top_income_sources: list[TopEntry]
    top_expense_by_count: list[TopEntry]
    top_expense_by_amount: list[TopEntry]
    top_grocery_by_count: list[TopEntry]
    top_grocery_by_amount: list[TopEntry]
    top_food: list[TopEntry]
    top_fast_foods: list[TopEntry]
    top_coffee_shops: list[TopEntry]

    monthly: list[MonthlyPoint]
    expense_by_category: list[tuple[str, float]]

    special_cases: dict[str, float | int]
    tagged_mentions: list[TaggedMention]
    shopee_orders: list[str]
    lazada_orders: list[str]
    amazon_orders: list[str]


def _matches_any(name: str, needles: Iterable[str]) -> bool:
    return any(
        re.search(rf'\b{re.escape(n)}\b', name, re.IGNORECASE) for n in needles
    )


def _top_n_food(food_data: dict, needles: Iterable[str], n: int = 10,
                require_coffee_flag: bool = False) -> list[TopEntry]:
    filtered = []
    for name, data in food_data.items():
        count, total, first, is_coffee = data
        if require_coffee_flag and not is_coffee:
            continue
        if _matches_any(name, needles):
            filtered.append(TopEntry(name, count, total, first))
    filtered.sort(key=lambda e: e.count, reverse=True)
    return filtered[:n]


def analyze(xlsx_bytes: bytes) -> AnalysisResult:
    """Run the full analysis against an in-memory xlsx file."""
    workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(values_only=True))

    try:
        idx = {col: header_row.index(col) for col in EXPECTED_COLUMNS}
    except ValueError as exc:
        missing = [c for c in EXPECTED_COLUMNS if c not in header_row]
        raise ValueError(
            f"Money Manager xlsx is missing expected columns: {missing}. "
            f"Required columns: {EXPECTED_COLUMNS}"
        ) from exc

    total_expense = 0.0
    total_expense_count = 0
    total_income = 0.0
    total_income_count = 0

    expense_account_data: dict[str, int] = {}
    income_from_data: dict[str, tuple[int, float, str]] = {}
    purchase_from_data: dict[str, tuple[int, float, str]] = {}
    food_data: dict[str, tuple[int, float, str, bool]] = {}
    grocery_data: dict[str, tuple[int, float, str]] = {}

    monthly_data: dict[str, dict[str, float]] = {}
    expense_by_category: dict[str, float] = {}

    total_shopee_count = 0
    shopee_orders: list[str] = []
    total_lazada_count = 0
    lazada_orders: list[str] = []
    total_amazon_count = 0
    amazon_orders: list[str] = []
    total_grab_food_count = 0
    total_grab_car_count = 0
    total_foodpanda_count = 0
    total_711 = 0

    # @Name tag aggregation. Keyed by lowercased name for dedupe; value tracks
    # the original-cased name first seen, the row count, and the summed amount.
    # Semantic matches the original Christian/Shane logic: each mention adds
    # the full row amount (no per-name splitting).
    tagged: dict[str, dict] = {}
    tag_pattern = re.compile(r'@(\w+)')

    for row in sheet.iter_rows(min_row=2, values_only=True):
        period_value = row[idx["Period"]]
        account_value = row[idx["Accounts"]]
        category_value = row[idx["Category"]]
        subcategory_value = row[idx["Subcategory"]]
        amount_value = row[idx["Amount"]]
        income_expense_value = row[idx["Income/Expense"]]
        note_value = row[idx["Note"]]
        description_value = row[idx["Description"]]

        if not isinstance(amount_value, (int, float)):
            continue
        if not isinstance(period_value, datetime):
            continue

        month_key = period_value.strftime("%Y-%m")
        month_label = period_value.strftime("%b %Y")
        bucket = monthly_data.setdefault(
            month_key, {"label": month_label, "income": 0.0, "expense": 0.0}
        )

        if income_expense_value == "Exp.":
            total_expense_count += 1
            total_expense += amount_value
            bucket["expense"] = round(bucket["expense"] + amount_value, 2)

            if account_value:
                expense_account_data[account_value] = expense_account_data.get(account_value, 0) + 1

            if category_value:
                expense_by_category[category_value] = round(
                    expense_by_category.get(category_value, 0.0) + amount_value, 2
                )

            if note_value:
                if note_value not in purchase_from_data:
                    purchase_from_data[note_value] = (0, 0.0, period_value.strftime("%B %d, %Y"))
                count, amt_total, first = purchase_from_data[note_value]
                if datetime.strptime(first, "%B %d, %Y") > period_value:
                    first = period_value.strftime("%B %d, %Y")
                purchase_from_data[note_value] = (
                    count + 1, round(amt_total + amount_value, 2), first
                )

                if category_value == "Food":
                    is_coffee_shop = subcategory_value in ("Cafe Hopping", "Grab", "Foodpanda")
                    if note_value not in food_data:
                        food_data[note_value] = (0, 0.0, period_value.strftime("%B %d, %Y"), is_coffee_shop)
                    fcount, famt, ffirst, _ = food_data[note_value]
                    if datetime.strptime(ffirst, "%B %d, %Y") > period_value:
                        ffirst = period_value.strftime("%B %d, %Y")
                    food_data[note_value] = (
                        fcount + 1, round(famt + amount_value, 2), ffirst, is_coffee_shop
                    )
                    if subcategory_value == "Grab":
                        total_grab_food_count += 1
                    if subcategory_value == "Foodpanda":
                        total_foodpanda_count += 1

                elif category_value == "Grocery":
                    if note_value not in grocery_data:
                        grocery_data[note_value] = (0, 0.0, period_value.strftime("%B %d, %Y"))
                    gcount, gamt, gfirst = grocery_data[note_value]
                    if datetime.strptime(gfirst, "%B %d, %Y") > period_value:
                        gfirst = period_value.strftime("%B %d, %Y")
                    grocery_data[note_value] = (
                        gcount + 1, round(gamt + amount_value, 2), gfirst
                    )

                elif category_value == "Transportation":
                    if re.search(r'^Grab', note_value, re.IGNORECASE):
                        total_grab_car_count += 1

                if re.search(r'Shopee$', note_value, re.IGNORECASE):
                    total_shopee_count += 1
                    if description_value:
                        item = re.sub(r'@Christian|@Shane', '', description_value).strip()
                        if item:
                            shopee_orders.append(item)
                if re.search(r'Lazada$', note_value, re.IGNORECASE):
                    total_lazada_count += 1
                    if description_value:
                        item = re.sub(r'@Christian|@Shane', '', description_value).strip()
                        if item:
                            lazada_orders.append(item)
                if re.search(r'Amazon$', note_value, re.IGNORECASE):
                    total_amazon_count += 1
                    if description_value:
                        item = re.sub(r'@Christian|@Shane', '', description_value).strip()
                        if item:
                            amazon_orders.append(item)
                if re.search(r'^711', note_value, re.IGNORECASE):
                    total_711 += 1

                if description_value:
                    # Dedupe within a single description so @Shane @shane
                    # doesn't double-count this row.
                    seen_in_row = set()
                    for match in tag_pattern.finditer(description_value):
                        raw = match.group(1)
                        key = raw.lower()
                        if key in seen_in_row:
                            continue
                        seen_in_row.add(key)
                        bucket = tagged.setdefault(
                            key, {"name": raw, "count": 0, "amount": 0.0}
                        )
                        bucket["count"] += 1
                        bucket["amount"] = round(bucket["amount"] + amount_value, 2)

        elif income_expense_value == "Income":
            total_income_count += 1
            total_income += amount_value
            bucket["income"] = round(bucket["income"] + amount_value, 2)

            if note_value:
                if note_value not in income_from_data:
                    income_from_data[note_value] = (0, 0.0, period_value.strftime("%B %d, %Y"))
                count, amt_total, first = income_from_data[note_value]
                if datetime.strptime(first, "%B %d, %Y") > period_value:
                    first = period_value.strftime("%B %d, %Y")
                income_from_data[note_value] = (
                    count + 1, round(amt_total + amount_value, 2), first
                )

    monthly = [
        MonthlyPoint(month=k, label=v["label"], income=v["income"], expense=v["expense"])
        for k, v in sorted(monthly_data.items())
    ]

    expense_accounts = sorted(expense_account_data.items(), key=lambda x: x[1], reverse=True)

    def _to_entries_by_count(data: dict, limit: int) -> list[TopEntry]:
        items = sorted(data.items(), key=lambda x: x[1][0], reverse=True)[:limit]
        return [TopEntry(name, d[0], d[1], d[2]) for name, d in items]

    def _to_entries_by_amount(data: dict, limit: int) -> list[TopEntry]:
        items = sorted(data.items(), key=lambda x: x[1][1], reverse=True)[:limit]
        return [TopEntry(name, d[0], d[1], d[2]) for name, d in items]

    top_income_sources = [
        TopEntry(name, d[0], d[1], d[2])
        for name, d in sorted(income_from_data.items(), key=lambda x: x[1][1], reverse=True)[:10]
    ]

    top_expense_by_count = _to_entries_by_count(purchase_from_data, 30)
    top_expense_by_amount = _to_entries_by_amount(purchase_from_data, 10)
    top_grocery_by_count = _to_entries_by_count(grocery_data, 10)
    top_grocery_by_amount = _to_entries_by_amount(grocery_data, 10)

    top_food = [
        TopEntry(name, d[0], d[1], d[2])
        for name, d in sorted(food_data.items(), key=lambda x: x[1][0], reverse=True)[:30]
    ]
    top_fast_foods = _top_n_food(food_data, FAST_FOODS, n=10)
    top_coffee_shops = _top_n_food(food_data, COFFEE_SHOPS, n=10, require_coffee_flag=True)

    return AnalysisResult(
        total_income=round(total_income, 2),
        total_expense=round(total_expense, 2),
        balance=round(total_income - total_expense, 2),
        total_income_count=total_income_count,
        total_expense_count=total_expense_count,
        expense_accounts=expense_accounts,
        top_income_sources=top_income_sources,
        top_expense_by_count=top_expense_by_count,
        top_expense_by_amount=top_expense_by_amount,
        top_grocery_by_count=top_grocery_by_count,
        top_grocery_by_amount=top_grocery_by_amount,
        top_food=top_food,
        top_fast_foods=top_fast_foods,
        top_coffee_shops=top_coffee_shops,
        monthly=monthly,
        expense_by_category=sorted(expense_by_category.items(), key=lambda x: x[1], reverse=True),
        special_cases={
            "shopee_count": total_shopee_count,
            "lazada_count": total_lazada_count,
            "amazon_count": total_amazon_count,
            "foodpanda_count": total_foodpanda_count,
            "grab_food_count": total_grab_food_count,
            "grab_car_count": total_grab_car_count,
            "seven_eleven_count": total_711,
        },
        tagged_mentions=sorted(
            (
                TaggedMention(name=v["name"], count=v["count"], amount=v["amount"])
                for v in tagged.values()
            ),
            key=lambda t: t.amount,
            reverse=True,
        ),
        shopee_orders=shopee_orders,
        lazada_orders=lazada_orders,
        amazon_orders=amazon_orders,
    )
